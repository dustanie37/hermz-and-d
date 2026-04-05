#!/usr/bin/env python3
"""
Hermz & D — Oscar Data Import Script
======================================
Reads 'Hermz and D Oscar Picks.xlsx' and generates a SQL file to populate
the Supabase database with all Oscar data (2008–2026).

Run:  python3 hermz_oscar_import.py
Output: supabase/oscar_import.sql
"""

import openpyxl
import re
from datetime import datetime, time

XLSX_PATH = '/sessions/amazing-bold-planck/mnt/Hermz and D Movies/Hermz and D Oscar Picks.xlsx'
OUTPUT_SQL = '/sessions/amazing-bold-planck/mnt/Hermz and D Movies/hermz-and-d/supabase/oscar_import.sql'

# ---------------------------------------------------------------------------
# Category alias map: spreadsheet labels → canonical DB category names
# ---------------------------------------------------------------------------
CATEGORY_ALIAS_MAP = {
    'Best Picture':              'Best Picture',
    'Best Director':             'Best Director',
    'Best Actor':                'Best Actor',
    'Best Actress':              'Best Actress',
    'Best Supporting Actor':     'Best Supporting Actor',
    'Best Supporting Actress':   'Best Supporting Actress',
    'Original Screenplay':       'Best Original Screenplay',
    'Adapted Screenplay':        'Best Adapted Screenplay',
    'Adapated Screenplay':       'Best Adapted Screenplay',   # typo in spreadsheet
    'Adaped Screenplay':         'Best Adapted Screenplay',   # alternate typo
    'Animated Feature':          'Best Animated Feature Film',
    'Animated Feature ':         'Best Animated Feature Film', # trailing space
    'Foreign Language':          'Best International Feature Film',
    'International Film':        'Best International Feature Film',
    'International Feature':     'Best International Feature Film',
    'Art Direction':             'Best Production Design',
    'Production Design':         'Best Production Design',
    'Cinematography':            'Best Cinematography',
    'Costume Design':            'Best Costume Design',
    'Doc Feature':               'Best Documentary Feature Film',
    'Documentary Feature':       'Best Documentary Feature Film',
    'Doc Short':                 'Best Documentary Short Film',
    'Documentary Short':         'Best Documentary Short Film',
    'Film Editing':              'Best Film Editing',
    'Makeup':                    'Best Makeup and Hairstyling',
    'Makeup & Hairstyling':      'Best Makeup and Hairstyling',
    'Makeup/Hair':               'Best Makeup and Hairstyling',
    'Makeup & Hair':             'Best Makeup and Hairstyling',
    'Visual Effects':            'Best Visual Effects',
    'Original Score':            'Best Original Score',
    'Original Song':             'Best Original Song',
    'Animated Short':            'Best Animated Short Film',
    'Live Action Short':         'Best Live Action Short Film',
    'Sound Editing':             'Best Sound Editing',
    'Sound Mixing':              'Best Sound Mixing',
    'Sound Design':              'Best Sound Mixing',  # 2008 spreadsheet label
    'Best Sound':                'Best Sound',
    'Casting':                   'Best Casting',
}

ALIAS_KEYS = set(CATEGORY_ALIAS_MAP.keys())

# ---------------------------------------------------------------------------
# Year winners (from references.md, verified against spreadsheet scores)
# ---------------------------------------------------------------------------
YEAR_WINNERS = {
    2008: 'matt',   2009: 'dustin', 2010: 'matt',   2011: 'dustin',
    2012: 'dustin', 2013: 'matt',   2014: 'dustin', 2015: 'dustin',
    2016: 'matt',   2017: 'matt',   2018: 'dustin', 2019: 'matt',
    2020: 'dustin', 2021: 'matt',   2022: 'dustin', 2023: 'dustin',
    2024: 'matt',   2025: 'dustin', 2026: 'dustin',
}
TIEBREAKER_YEARS = {2010, 2011, 2018, 2026}

# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------
def sql_str(s):
    """Escape a value as a SQL string literal, or NULL."""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

def sql_date(s):
    """Return a SQL date literal or NULL."""
    if s is None:
        return 'NULL'
    return f"'{s}'"

def sql_interval(t):
    """Convert datetime.time or various string formats to a PostgreSQL INTERVAL."""
    if t is None:
        return 'NULL'
    if isinstance(t, time):
        return f"'{t.hour:02d}:{t.minute:02d}:{t.second:02d}'::INTERVAL"
    if isinstance(t, str):
        t = t.strip()
        # '13 min 8 secs' style (monologue)
        m = re.match(r'(\d+)\s*min\w*\s+(\d+)\s*sec', t, re.IGNORECASE)
        if m:
            mins, secs = int(m.group(1)), int(m.group(2))
            return f"'00:{mins:02d}:{secs:02d}'::INTERVAL"
        # 'H:MM' or 'HH:MM' style (2009 tiebreaker)
        m = re.match(r'^(\d{1,2}):(\d{2})$', t)
        if m:
            hrs, mins = int(m.group(1)), int(m.group(2))
            return f"'{hrs:02d}:{mins:02d}:00'::INTERVAL"
    return 'NULL'

def sql_bool(b):
    return 'TRUE' if b else 'FALSE'

# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------
def parse_ceremony_date(title):
    """Extract 'YYYY-MM-DD' from ceremony title (e.g. '98th Academy Awards - March 15, 2026')."""
    if not title:
        return None
    m = re.search(r'(\w+ \d+,?\s*\d{4})', title)
    if m:
        raw = m.group(1).replace(',', '')
        for fmt in ('%B %d %Y', '%B %d, %Y'):
            try:
                return datetime.strptime(raw, fmt.replace(',', '')).strftime('%Y-%m-%d')
            except ValueError:
                pass
    return None

def parse_sheet(ws, year):
    """Parse one year sheet. Returns a dict with ceremony info, categories, nominees."""
    rows = list(ws.iter_rows(values_only=True))

    # ---- Ceremony title (row 1, any column) --------------------------------
    ceremony_title = None
    ceremony_date = None
    for cell in rows[0]:
        if cell and isinstance(cell, str) and 'Academy Awards' in cell:
            ceremony_title = cell.strip()
            ceremony_date = parse_ceremony_date(ceremony_title)
            break

    # ---- Category rows (cols A–F) ------------------------------------------
    categories = {}  # canonical_name → {winner, matt_guess, dustin_guess, matt_correct, dustin_correct}
    tiebreaker = {}
    monologue = {}
    total_matt = None
    total_dustin = None

    for row in rows:
        if not any(row):
            continue
        col_a = row[0]
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None
        col_e = row[4] if len(row) > 4 else None
        col_f = row[5] if len(row) > 5 else None

        # Tiebreaker rows
        if isinstance(col_a, str) and col_a.strip().lower() == 'run time tiebreaker':
            tiebreaker = {'actual': col_b, 'matt': col_c, 'dustin': col_d}
            continue
        if isinstance(col_a, str) and 'monologue' in col_a.lower():
            monologue = {'actual': col_b, 'matt': col_c, 'dustin': col_d}
            continue

        # Score summary row (TOTAL or CORRECT)
        if col_b in ('TOTAL', 'CORRECT'):
            # col_e/col_f always hold integer counts when present
            if isinstance(col_e, (int, float)):
                total_matt = int(col_e)
            elif isinstance(col_c, (int, float)):
                total_matt = int(col_c)
            if isinstance(col_f, (int, float)):
                total_dustin = int(col_f)
            elif isinstance(col_d, (int, float)):
                total_dustin = int(col_d)
            continue

        # Category data row
        if isinstance(col_a, str) and col_a.strip() in ALIAS_KEYS:
            cat = CATEGORY_ALIAS_MAP[col_a.strip()]
            categories[cat] = {
                'winner':         col_b,
                'matt_guess':     col_c,
                'dustin_guess':   col_d,
                'matt_correct':   col_e == 1,
                'dustin_correct': col_f == 1,
            }

    # ---- Nominees (cols G–M, indices 6–12) ---------------------------------
    nominees = {}  # canonical_category → [nominee_name, ...]

    for col_idx in range(6, 13):   # cols G–M (0-indexed from row tuple)
        current_cat = None
        current_noms = []

        def flush():
            nonlocal current_noms
            if current_cat and current_noms:
                nominees[current_cat] = current_noms[:]
            current_noms = []

        for row in rows:
            if len(row) <= col_idx:
                continue
            cell = row[col_idx]

            if cell is None:
                flush()
                current_cat = None
            elif isinstance(cell, str) and cell.strip() in ALIAS_KEYS:
                flush()
                current_cat = CATEGORY_ALIAS_MAP[cell.strip()]
                current_noms = []
            elif isinstance(cell, (int, float)):
                # Stray numeric value in nominee block — skip
                pass
            else:
                if current_cat:
                    val = str(cell).strip()
                    if val:
                        current_noms.append(val)

        flush()   # end of column

    return {
        'year':            year,
        'ceremony_title':  ceremony_title,
        'ceremony_date':   ceremony_date,
        'categories':      categories,
        'nominees':        nominees,
        'tiebreaker':      tiebreaker,
        'monologue':       monologue,
        'total_matt':      total_matt,
        'total_dustin':    total_dustin,
    }

# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------
def generate_sql(all_data):
    lines = []
    lines.append("-- =============================================================")
    lines.append("-- HERMZ & D — Oscar Data Import")
    lines.append("-- Generated by hermz_oscar_import.py")
    lines.append("-- Run this in the Supabase SQL editor AFTER schema.sql")
    lines.append("-- =============================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # ---- 1. oscar_years -----------------------------------------------------
    lines.append("-- -------------------------------------------------------------")
    lines.append("-- 1. OSCAR YEARS (19 ceremonies: 2008–2026)")
    lines.append("-- -------------------------------------------------------------")
    lines.append("")

    for year in sorted(all_data.keys()):
        d = all_data[year]
        tb = d['tiebreaker']
        ml = d['monologue']
        winner = YEAR_WINNERS.get(year, 'pending')
        tb_used = year in TIEBREAKER_YEARS

        lines.append(f"INSERT INTO public.oscar_years")
        lines.append(f"  (year, ceremony_name, ceremony_date,")
        lines.append(f"   actual_runtime, matt_runtime_guess, dustin_runtime_guess,")
        lines.append(f"   actual_monologue, matt_monologue_guess, dustin_monologue_guess,")
        lines.append(f"   winner, tiebreaker_used)")
        lines.append(f"VALUES")
        lines.append(f"  ({year},")
        lines.append(f"   {sql_str(d['ceremony_title'])},")
        lines.append(f"   {sql_date(d['ceremony_date'])},")
        lines.append(f"   {sql_interval(tb.get('actual'))},")
        lines.append(f"   {sql_interval(tb.get('matt'))},")
        lines.append(f"   {sql_interval(tb.get('dustin'))},")
        lines.append(f"   {sql_interval(ml.get('actual') if ml else None)},")
        lines.append(f"   {sql_interval(ml.get('matt') if ml else None)},")
        lines.append(f"   {sql_interval(ml.get('dustin') if ml else None)},")
        lines.append(f"   '{winner}', {sql_bool(tb_used)}")
        lines.append(f"  );")
        lines.append("")

    # ---- 2. oscar_nominees --------------------------------------------------
    lines.append("-- -------------------------------------------------------------")
    lines.append("-- 2. OSCAR NOMINEES")
    lines.append("-- -------------------------------------------------------------")
    lines.append("")

    for year in sorted(all_data.keys()):
        d = all_data[year]
        lines.append(f"-- {year} nominees")
        for cat_name, nom_list in d['nominees'].items():
            cat_data = d['categories'].get(cat_name, {})
            actual_winner = cat_data.get('winner')
            for i, nom in enumerate(nom_list, 1):
                is_win = (actual_winner and
                          nom.lower().strip() == str(actual_winner).lower().strip())
                lines.append(
                    f"INSERT INTO public.oscar_nominees "
                    f"(year_id, category_id, nominee_name, is_winner, display_order)"
                )
                lines.append(f"SELECT")
                lines.append(f"  (SELECT id FROM public.oscar_years WHERE year = {year}),")
                lines.append(f"  (SELECT id FROM public.oscar_categories WHERE name = {sql_str(cat_name)}),")
                lines.append(f"  {sql_str(nom)},")
                lines.append(f"  {sql_bool(is_win)},")
                lines.append(f"  {i};")
        lines.append("")

    # ---- 3. oscar_guesses ---------------------------------------------------
    lines.append("-- -------------------------------------------------------------")
    lines.append("-- 3. OSCAR GUESSES (Matt and Dustin, all years)")
    lines.append("-- -------------------------------------------------------------")
    lines.append("")

    for year in sorted(all_data.keys()):
        d = all_data[year]
        lines.append(f"-- {year} guesses")
        for cat_name, cat_data in d['categories'].items():
            matt_guess   = cat_data.get('matt_guess')
            dustin_guess = cat_data.get('dustin_guess')
            matt_correct   = cat_data.get('matt_correct')
            dustin_correct = cat_data.get('dustin_correct')

            # Matt
            if matt_guess:
                lines.append(
                    f"INSERT INTO public.oscar_guesses "
                    f"(year_id, category_id, user_id, guess, is_correct, locked)"
                )
                lines.append(f"SELECT")
                lines.append(f"  (SELECT id FROM public.oscar_years WHERE year = {year}),")
                lines.append(f"  (SELECT id FROM public.oscar_categories WHERE name = {sql_str(cat_name)}),")
                lines.append(f"  (SELECT id FROM public.profiles WHERE username = 'matt'),")
                lines.append(f"  {sql_str(matt_guess)},")
                lines.append(f"  {sql_bool(matt_correct)},")
                lines.append(f"  TRUE;")

            # Dustin
            if dustin_guess:
                lines.append(
                    f"INSERT INTO public.oscar_guesses "
                    f"(year_id, category_id, user_id, guess, is_correct, locked)"
                )
                lines.append(f"SELECT")
                lines.append(f"  (SELECT id FROM public.oscar_years WHERE year = {year}),")
                lines.append(f"  (SELECT id FROM public.oscar_categories WHERE name = {sql_str(cat_name)}),")
                lines.append(f"  (SELECT id FROM public.profiles WHERE username = 'dustin'),")
                lines.append(f"  {sql_str(dustin_guess)},")
                lines.append(f"  {sql_bool(dustin_correct)},")
                lines.append(f"  TRUE;")

        lines.append("")

    lines.append("COMMIT;")
    lines.append("")
    lines.append("-- =============================================================")
    lines.append("-- END OF IMPORT — verify with:")
    lines.append("-- SELECT * FROM public.v_oscar_year_summary ORDER BY year;")
    lines.append("-- =============================================================")

    return "\n".join(lines)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    print("Loading workbook...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    year_sheets = [str(y) for y in range(2008, 2027) if str(y) in wb.sheetnames]
    print(f"Found year sheets: {year_sheets}")

    all_data = {}
    for yr_str in year_sheets:
        ws = wb[yr_str]
        data = parse_sheet(ws, int(yr_str))
        all_data[int(yr_str)] = data

    # ---- Validation summary -------------------------------------------------
    print("\n=== VALIDATION SUMMARY ===")
    print(f"{'Year':<6} {'Ceremony':<45} {'Cats':>4} {'Noms':>4} {'Matt':>4} {'Dust':>4}")
    print("-" * 75)

    all_ok = True
    for year in sorted(all_data.keys()):
        d = all_data[year]
        cats = len(d['categories'])
        noms = sum(len(v) for v in d['nominees'].values())
        matt = d['total_matt']
        dust = d['total_dustin']

        # Expected category counts
        if year <= 2020:
            expected_cats = 24
        elif year <= 2025:
            expected_cats = 23
        else:
            expected_cats = 24

        flag = '' if cats == expected_cats else f' ← EXPECTED {expected_cats}'
        print(f"{year:<6} {(d['ceremony_title'] or 'UNKNOWN')[:44]:<45} {cats:>4} {noms:>4} {(matt or '?'):>4} {(dust or '?'):>4}{flag}")
        if cats != expected_cats:
            all_ok = False

    print()

    # Verify scores against references
    from_refs = {
        2008:(15,13), 2009:(17,20), 2010:(16,16), 2011:(17,17),
        2012:(16,18), 2013:(17,16), 2014:(21,22), 2015:(17,20),
        2016:(19,18), 2017:(17,14), 2018:(20,20), 2019:(17,14),
        2020:(19,20), 2021:(17,16), 2022:(16,19), 2023:(17,18),
        2024:(19,16), 2025:(16,18), 2026:(20,20),
    }

    print("=== SCORE VERIFICATION (vs references.md) ===")
    score_ok = True
    for year in sorted(all_data.keys()):
        d = all_data[year]
        expected_m, expected_d = from_refs[year]
        actual_m, actual_d = d['total_matt'], d['total_dustin']
        status = '✓' if (actual_m == expected_m and actual_d == expected_d) else '✗ MISMATCH'
        if status != '✓':
            score_ok = False
        print(f"  {year}: Matt={actual_m}/{expected_m}  Dustin={actual_d}/{expected_d}  {status}")

    print()
    if all_ok and score_ok:
        print("✓ All validations passed. Generating SQL...")
    else:
        print("⚠ Some validations failed — review above before importing.")

    # ---- Generate SQL -------------------------------------------------------
    sql = generate_sql(all_data)
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write(sql)

    line_count = sql.count('\n')
    print(f"\nSQL written to: {OUTPUT_SQL}")
    print(f"Lines: {line_count}")
    print("Done!")
