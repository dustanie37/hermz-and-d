#!/usr/bin/env python3
"""
Hermz & D — Movie Rankings Import Script
=========================================
Parses both movie ranking XLSX files and generates movie_import.sql.

Usage:
    python3 movie_import.py

Output:
    movie_import.sql   — INSERT statements for films, individual_rankings,
                         and combined_rankings tables.
                         Run in Supabase SQL Editor AFTER movie_schema_update.sql.

Sources:
    2001 and 2007 Hermz and D Top 100 Favorite Films.xlsx
    2016 and 2026 Hermz and D Top 100 Favorite Movies.xlsx
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
WORKSPACE  = SCRIPT_DIR.parent.parent   # workspace root (contains both xlsx files)

FILE_0107 = WORKSPACE / "2001 and 2007 Hermz and D Top 100 Favorite Films.xlsx"
FILE_1626 = WORKSPACE / "2016 and 2026 Hermz and D Top 100 Favorite Movies.xlsx"
OUTPUT    = SCRIPT_DIR / "movie_import.sql"

# ── Year normalisation ────────────────────────────────────────────────────────
# Films that appear with inconsistent release years across sheets.
# Keys are lowercase stripped title (no punctuation), values are canonical year.
YEAR_NORMALIZE = {
    "saving private ryan":                          1998,
    "memento":                                      2000,
    "life is beautiful":                            1997,
    "la vita e bella":                              1997,
    "life is beautiful la vita e bella":            1997,
    "spaceballs":                                   1987,
    "superman ii":                                  1980,
    "the pink panther strikes again":               1976,
    "othello":                                      1995,   # Branagh version
    "300":                                          2006,
    "1917":                                         2019,
    "noises off":                                   1992,
    "noises off!":                                  1992,
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def sql_str(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"

def sql_int(v):
    if v is None or (isinstance(v, str) and v.strip().upper() in ('NR', '#NAME?', '-', '')):
        return "NULL"
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return "NULL"

def sql_bool(v):
    if v is None or v == '-':
        return "FALSE"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if str(v).strip().upper() in ('X', 'TRUE', '1', 'YES'):
        return "TRUE"
    return "FALSE"

def sql_score(v, min_=1, max_=10):
    if v is None:
        return "NULL"
    try:
        n = int(float(v))
        if min_ <= n <= max_:
            return str(n)
        return "NULL"
    except (TypeError, ValueError):
        return "NULL"

def is_rank(v):
    if v is None or isinstance(v, bool):
        return False
    if isinstance(v, str) and v.strip().upper() in ('NR', '#NAME?', '-', ''):
        return False
    try:
        int(float(v))
        return True
    except (TypeError, ValueError):
        return False

def normalize_title(title):
    """Return a clean display title (preserves capitalisation)."""
    if title is None:
        return None
    # Convert floats/ints that Excel stored as numbers (e.g. 1917.0, 300.0)
    if isinstance(title, float):
        title = str(int(title))
    elif isinstance(title, int):
        title = str(title)
    t = str(title).strip()
    t = t.replace('\u2019', "'").replace('\u2018', "'")
    t = t.replace('\u201c', '"').replace('\u201d', '"')
    t = t.replace('\u2013', '-').replace('\u2014', '-')
    t = re.sub(r'\s+', ' ', t)
    return t or None

def bare_key(title):
    """Lowercase, punctuation-free key for matching (no year)."""
    t = normalize_title(title)
    if not t:
        return None
    k = re.sub(r"[^a-z0-9 ]", '', t.lower()).strip()
    return re.sub(r'\s+', ' ', k)

def canonical_year(title, raw_year):
    """Return the normalised year for a film, applying known overrides."""
    bk = bare_key(title)
    if bk and bk in YEAR_NORMALIZE:
        return YEAR_NORMALIZE[bk]
    if raw_year is None:
        return None
    try:
        return int(float(raw_year))
    except (TypeError, ValueError):
        return None

def film_key(title, year=None):
    """Lookup key: bare_title + optional :year."""
    bk = bare_key(title)
    if bk is None:
        return None
    if year:
        return f"{bk}:{year}"
    return bk

# ── Film Registry ─────────────────────────────────────────────────────────────

films = {}           # key → film dict
film_counter = [1]

def _new_film(raw_title, year):
    fid = film_counter[0]
    film_counter[0] += 1
    return {
        'id': fid,
        'title': raw_title,
        'release_year': year,
        'director': None,
        'actor_1': None, 'actor_2': None, 'actor_3': None, 'actor_4': None,
        'custom_genre_1': None, 'custom_genre_2': None,
        'acclaim_score': None,
        'oscar_nominations': 0,
        'oscar_wins': 0,
        'won_best_picture': False, 'won_best_director': False,
        'won_best_actor': False, 'won_best_actress': False,
        'won_best_supp_actor': False, 'won_best_supp_actress': False,
        'won_screenplay': False, 'won_cinematography': False,
        'won_production_design': False,
        'afi_top100_rank': None, 'afi_comedies_rank': None,
        'imdb_top250_rank': None, 'nyt_2000s_rank': None,
        'sight_sound_2022_rank': None, 'variety_comedies_rank': None,
        'national_film_registry': False,
    }

def get_or_create_film(title, year=None):
    """Return film id; create if not yet seen."""
    raw_title = normalize_title(title)
    if not raw_title:
        return None
    yr = canonical_year(raw_title, year)
    key = film_key(raw_title, yr)
    if key not in films:
        # Also check if a bare-key (no year) entry already exists
        bk = bare_key(raw_title)
        existing = _find_by_bare_key(bk)
        if existing is not None:
            return existing
        films[key] = _new_film(raw_title, yr)
    return films[key]['id']

def get_film_id(title, year=None):
    """Look up an existing film id; returns None if not found."""
    raw_title = normalize_title(title)
    if not raw_title:
        return None
    yr = canonical_year(raw_title, year)

    # 1. Exact title+year match
    key = film_key(raw_title, yr)
    if key in films:
        return films[key]['id']

    # 2. Bare-key scan (ignore year differences)
    bk = bare_key(raw_title)
    return _find_by_bare_key(bk)

def _find_by_bare_key(bk):
    """Return id of the first film whose bare key starts with bk."""
    if not bk:
        return None
    for k, f in films.items():
        if k == bk or k.startswith(bk + ':'):
            return f['id']
    return None

# ── Individual Rankings ────────────────────────────────────────────────────────

individual_rankings = []
combined_rankings   = []

def add_individual_ranking(film_id, event_year, username, rank, total_score,
                            lead, supp, plot, dialogue, screenplay, direction,
                            cinematography, art_direction, influence, acclaim,
                            personal_impact, tens, nines, eights):
    individual_rankings.append({
        'film_id':               film_id,
        'event_year':            event_year,
        'username':              username,
        'rank':                  rank,
        'total_score':           total_score,
        'score_lead_performance':lead,
        'score_supp_performance':supp,
        'score_plot':            plot,
        'score_dialogue':        dialogue,
        'score_screenplay':      screenplay,
        'score_direction':       direction,
        'score_cinematography':  cinematography,
        'score_production_design':art_direction,
        'score_influence':       influence,
        'score_acclaim':         acclaim,
        'score_personal_impact': personal_impact,
        'tb_tens':               tens,
        'tb_nines':              nines,
        'tb_eights':             eights,
    })

# ── 2001 individual lists ─────────────────────────────────────────────────────
# Cols: RANK(0) Title(1) Year(2) Lead(3) Sup(4) Plot(5) Dialogue(6)
#       Direction(7) Cinematog(8) Influence(9) Acclaim(10) Personal(11)
#       TOTAL(12) Tens(13) Nines(14) Eights(15)

def parse_2001_list(ws, username):
    for row in ws.iter_rows(values_only=True):
        if not is_rank(row[0]):
            continue
        title = normalize_title(row[1])
        if not title:
            continue
        fid = get_or_create_film(title, row[2])
        add_individual_ranking(
            film_id=fid, event_year=2001, username=username,
            rank=int(float(row[0])),
            total_score=int(float(row[12])) if row[12] else None,
            lead=row[3], supp=row[4],
            plot=row[5], dialogue=row[6],
            screenplay=None, direction=row[7],
            cinematography=row[8], art_direction=None,
            influence=row[9], acclaim=row[10],
            personal_impact=row[11],
            tens=row[13], nines=row[14], eights=row[15],
        )

# ── 2007 individual lists ─────────────────────────────────────────────────────
# Header rows 0–3; data row 4+
# Cols: 2007rank(0) 2001rank(1) Change(2) Title(3) Year(4)
#       Lead(5) Supp(6) Screenplay(7) Direction(8) Cinematog(9) ArtDir(10)
#       Influence(11) Acclaim(12) Personal(13) TOTAL(14) Tens(15) Nines(16) Eights(17)

def parse_2007_list(ws, username):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4 or not is_rank(row[0]):
            continue
        title = normalize_title(row[3])
        if not title:
            continue
        fid = get_or_create_film(title, row[4])
        add_individual_ranking(
            film_id=fid, event_year=2007, username=username,
            rank=int(float(row[0])),
            total_score=int(float(row[14])) if row[14] else None,
            lead=row[5], supp=row[6],
            plot=None, dialogue=None,
            screenplay=row[7], direction=row[8],
            cinematography=row[9], art_direction=row[10],
            influence=row[11], acclaim=row[12],
            personal_impact=row[13],
            tens=row[15], nines=row[16], eights=row[17],
        )

# ── 2016 individual lists ─────────────────────────────────────────────────────
# Header rows 0–2; data row 3+
# Cols: 2001rank(0) 2007rank(1) 2016rank(2) Change01(3) Change07(4)
#       Title(5) Year(6) Total(7) Lead(8) Supp(9) Screenplay(10)
#       Direction(11) Cinematog(12) ArtDir(13) Influence(14) Acclaim(15)
#       Personal(16) Tens(17) Nines(18) Eights(19)

def parse_2016_list(ws, username):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not is_rank(row[2]):
            continue
        title = normalize_title(row[5])
        if not title:
            continue
        fid = get_or_create_film(title, row[6])
        add_individual_ranking(
            film_id=fid, event_year=2016, username=username,
            rank=int(float(row[2])),
            total_score=int(float(row[7])) if row[7] else None,
            lead=row[8], supp=row[9],
            plot=None, dialogue=None,
            screenplay=row[10], direction=row[11],
            cinematography=row[12], art_direction=row[13],
            influence=row[14], acclaim=row[15],
            personal_impact=row[16],
            tens=row[17], nines=row[18], eights=row[19],
        )

# ── 2026 individual lists ─────────────────────────────────────────────────────
# Header rows 0–2; data row 3+
# Cols: 2001(0) 2007(1) 2016(2) 2026rank(3) Changes(4-7)
#       Title(8) Year(9) Total(10) Lead(11) Supp(12) Screenplay(13)
#       Direction(14) Cinematog(15) ProdDesign(16) Influence(17)
#       Acclaim(18) Personal(19) Tens(20) Nines(21) Eights(22)

def parse_2026_list(ws, username):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not is_rank(row[3]):
            continue
        title = normalize_title(row[8])
        if not title:
            continue
        fid = get_or_create_film(title, row[9])
        add_individual_ranking(
            film_id=fid, event_year=2026, username=username,
            rank=int(float(row[3])),
            total_score=int(float(row[10])) if row[10] else None,
            lead=row[11], supp=row[12],
            plot=None, dialogue=None,
            screenplay=row[13], direction=row[14],
            cinematography=row[15], art_direction=row[16],
            influence=row[17], acclaim=row[18],
            personal_impact=row[19],
            tens=row[20], nines=row[21], eights=row[22],
        )

# ── Combined rankings helpers ──────────────────────────────────────────────────

def add_combined_ranking(film_id, event_year, combined_rank,
                          dustin_rank, matt_rank, avg_rank,
                          dustin_score, matt_score, total_score,
                          dustin_impact, matt_impact, total_impact, total_tens):
    combined_rankings.append({
        'film_id':       film_id,
        'event_year':    event_year,
        'combined_rank': combined_rank,
        'dustin_rank':   dustin_rank,
        'matt_rank':     matt_rank,
        'avg_rank':      avg_rank,
        'dustin_score':  dustin_score,
        'matt_score':    matt_score,
        'total_score':   total_score,
        'dustin_impact': dustin_impact,
        'matt_impact':   matt_impact,
        'total_impact':  total_impact,
        'total_tens':    total_tens,
    })

# ── 2001 Combined Rankings ────────────────────────────────────────────────────
# Header row 0; data row 1+
# Cols: Rank(0) Film(1) DRank(2) MRank(3) AvgRank(4) DScore(5) MScore(6)
#       TotalScore(7) DImpact(8) MImpact(9) TotalImpact(10) TotalTens(11)

def parse_2001_combined(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 1 or not is_rank(row[0]):
            continue
        title = normalize_title(row[1])
        if not title:
            continue
        fid = get_film_id(title)           # always found — individual lists parsed first
        if fid is None:
            fid = get_or_create_film(title)
        add_combined_ranking(
            film_id=fid, event_year=2001,
            combined_rank=int(float(row[0])),
            dustin_rank=row[2], matt_rank=row[3], avg_rank=row[4],
            dustin_score=row[5], matt_score=row[6], total_score=row[7],
            dustin_impact=row[8], matt_impact=row[9], total_impact=row[10],
            total_tens=row[11],
        )

# ── 2007 Combined Rankings ────────────────────────────────────────────────────
# Header rows 0–1; data row 2+
# Cols: 2001rank(0) Rank(1) Change(2) Film(3) DRank(4) MRank(5)
#       AvgRank(6) DScore(7) MScore(8) TotalScore(9) DImpact(10)
#       MImpact(11) TotalImpact(12) TotalTens(13)

def parse_2007_combined(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2 or not is_rank(row[1]):
            continue
        title = normalize_title(row[3])
        if not title:
            continue
        fid = get_film_id(title)
        if fid is None:
            fid = get_or_create_film(title)
        add_combined_ranking(
            film_id=fid, event_year=2007,
            combined_rank=int(float(row[1])),
            dustin_rank=row[4], matt_rank=row[5], avg_rank=row[6],
            dustin_score=row[7], matt_score=row[8], total_score=row[9],
            dustin_impact=row[10], matt_impact=row[11], total_impact=row[12],
            total_tens=row[13],
        )

# ── 2016 Combined Rankings (sheet "2016 Combined Top 25" — actually all 48) ──
# Header row 0; data row 1+
# Cols: 2007rank(0) 2016rank(1) Change(2) Film(3) Year(4) DRank(5) MRank(6)
#       AvgRank(7) DScore(8) MScore(9) TotalScore(10) DImpact(11) MImpact(12)
#       TotalImpact(13) Tens(14)

def parse_2016_combined(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 1 or not is_rank(row[1]):
            continue
        title = normalize_title(row[3])
        year  = row[4]
        if not title:
            continue
        fid = get_film_id(title, year)
        if fid is None:
            fid = get_film_id(title)
        if fid is None:
            fid = get_or_create_film(title, year)
        add_combined_ranking(
            film_id=fid, event_year=2016,
            combined_rank=int(float(row[1])),
            dustin_rank=row[5], matt_rank=row[6], avg_rank=row[7],
            dustin_score=row[8], matt_score=row[9], total_score=row[10],
            dustin_impact=row[11], matt_impact=row[12], total_impact=row[13],
            total_tens=row[14],
        )

# ── 2026 Combined Rankings ────────────────────────────────────────────────────
# Header rows 0–1; data row 2+
# Cols: 2001(0) 2007(1) 2016(2) 2026rank(3) Changes(4-6) Film(7)
#       Year(8) DRank(9) MRank(10) AvgRank(11) DScore(12) MScore(13)
#       TotalScore(14) DImpact(15) MImpact(16) TotalImpact(17) TotalTens(18)

def parse_2026_combined(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2 or not is_rank(row[3]):
            continue
        title = normalize_title(row[7])
        year  = row[8]
        if not title:
            continue
        fid = get_film_id(title, year)
        if fid is None:
            fid = get_film_id(title)
        if fid is None:
            fid = get_or_create_film(title, year)
        add_combined_ranking(
            film_id=fid, event_year=2026,
            combined_rank=int(float(row[3])),
            dustin_rank=row[9], matt_rank=row[10], avg_rank=row[11],
            dustin_score=row[12], matt_score=row[13], total_score=row[14],
            dustin_impact=row[15], matt_impact=row[16], total_impact=row[17],
            total_tens=row[18],
        )

# ── Movie Metadata (2026 Movie Data sheet) ─────────────────────────────────────
# Header rows 0–1; data row 2+
# Cols: 2026Nominee(0) Title(1) Year(2) Acclaim(3) D(4) Hermz(5)
#       AFI100(6) AFIComedies(7) IMDB250(8) NYT2000s(9) SightSound(10)
#       VarietyComedies(11) NatFilmReg(12) Noms(13) Wins(14) Special(15)
#       BestPicture(16) BestDir(17) BestActor(18) BestActress(19)
#       BestSuppActor(20) BestSuppActress(21) Screenplay(22) Cinematog(23)
#       ProdDesign(24) Genre1(25) Genre2(26) Director(27)
#       Actor1(28) Actor2(29) Actor3(30) Actor4(31)

def parse_movie_data(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        title = normalize_title(row[1])
        year  = row[2]
        if not title:
            continue

        # Enrich existing film record (or create if somehow missing)
        fid = get_film_id(title, year)
        if fid is None:
            fid = get_or_create_film(title, year)

        # Find the film record by id to update
        f = next((v for v in films.values() if v['id'] == fid), None)
        if f is None:
            continue

        def enrich(field, raw):
            """Only set if not already populated."""
            if raw and str(raw).strip() not in ('', '-', 'N/A'):
                f[field] = normalize_title(raw) if isinstance(raw, str) else raw

        enrich('director',  row[27])
        enrich('actor_1',   row[28])
        enrich('actor_2',   row[29])
        enrich('actor_3',   row[30])
        enrich('actor_4',   row[31])
        enrich('custom_genre_1', row[25])
        enrich('custom_genre_2', row[26])

        # Acclaim score
        if row[3] is not None:
            try:
                f['acclaim_score'] = int(float(row[3]))
            except (TypeError, ValueError):
                pass

        # Oscar data
        try:
            f['oscar_nominations'] = int(float(row[13])) if row[13] else 0
        except (TypeError, ValueError):
            pass
        try:
            f['oscar_wins'] = int(float(row[14])) if row[14] else 0
        except (TypeError, ValueError):
            pass

        # Oscar wins by category
        f['won_best_picture']      = str(row[16]).strip() == 'X'
        f['won_best_director']     = str(row[17]).strip() == 'X'
        f['won_best_actor']        = str(row[18]).strip() == 'X'
        f['won_best_actress']      = str(row[19]).strip() == 'X'
        f['won_best_supp_actor']   = str(row[20]).strip() == 'X'
        f['won_best_supp_actress'] = str(row[21]).strip() == 'X'
        f['won_screenplay']        = str(row[22]).strip() == 'X'
        f['won_cinematography']    = str(row[23]).strip() == 'X'
        f['won_production_design'] = str(row[24]).strip() == 'X'

        # External list appearances
        def list_rank(v):
            if v is None or str(v).strip() in ('-', ''):
                return None
            try:
                return int(float(v))
            except (TypeError, ValueError):
                return None

        f['afi_top100_rank']        = list_rank(row[6])
        f['afi_comedies_rank']      = list_rank(row[7])
        f['imdb_top250_rank']       = list_rank(row[8])
        f['nyt_2000s_rank']         = list_rank(row[9])
        f['sight_sound_2022_rank']  = list_rank(row[10])
        f['variety_comedies_rank']  = list_rank(row[11])
        if row[12] and str(row[12]).strip() == 'X':
            f['national_film_registry'] = True

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading workbooks…")
    wb1 = openpyxl.load_workbook(FILE_0107, read_only=True, data_only=True)
    wb2 = openpyxl.load_workbook(FILE_1626, read_only=True, data_only=True)

    # Step 1 — Individual lists (creates all film records with correct years)
    print("Parsing individual rankings…")
    parse_2001_list(wb1["Martin 2001 List"], "dustin")
    parse_2001_list(wb1["Hermz 2001 List"],  "matt")
    parse_2007_list(wb1["Martin 2007 List"], "dustin")
    parse_2007_list(wb1["Hermz 2007 List"],  "matt")
    parse_2016_list(wb2["Martin 2016 List"], "dustin")
    parse_2016_list(wb2["Hermz 2016 List"],  "matt")
    parse_2026_list(wb2["Martin 2026 List"], "dustin")
    parse_2026_list(wb2["Hermz 2026 List"],  "matt")

    # Step 2 — Combined rankings (looks up existing films by title)
    print("Parsing combined rankings…")
    parse_2001_combined(wb1["2001 Combined Rankings"])
    parse_2007_combined(wb1["2007 Combined Rankings"])
    parse_2016_combined(wb2["2016 Combined Top 25"])
    parse_2026_combined(wb2["2026 Combined Rankings"])

    # Step 3 — Enrich film metadata from Movie Data sheet
    print("Parsing movie metadata…")
    parse_movie_data(wb2["2026 Movie Data"])

    unique_films = sorted(films.values(), key=lambda f: f['id'])

    # Sanity check — flag any remaining NULL-year films
    null_year = [f for f in unique_films if f['release_year'] is None]
    if null_year:
        print(f"\n⚠  {len(null_year)} films with NULL year (will still import fine):")
        for f in null_year:
            print(f"   ID {f['id']}: {f['title']}")

    # Deduplicate individual_rankings: same film cannot appear twice for the same
    # person in the same event. This can happen when two slightly-different title
    # variants in a spreadsheet both normalise to the same film_id via bare-key
    # matching. Keep the entry with the lowest (best) rank number.
    ir_dedup = {}
    for r in individual_rankings:
        key = (r['film_id'], r['event_year'], r['username'])
        if key not in ir_dedup or r['rank'] < ir_dedup[key]['rank']:
            ir_dedup[key] = r
    deduped_ir = list(ir_dedup.values())
    removed = len(individual_rankings) - len(deduped_ir)
    if removed:
        print(f"\n⚠  Deduplicated {removed} individual ranking row(s) (same film/event/user — kept lowest rank).")
    individual_rankings[:] = deduped_ir

    print(f"\nSummary:")
    print(f"  Unique films:          {len(unique_films)}")
    print(f"  Individual rankings:   {len(individual_rankings)}")
    print(f"  Combined rankings:     {len(combined_rankings)}")

    # ── Generate SQL ──────────────────────────────────────────────────────────
    lines = []
    lines.append("-- ============================================================")
    lines.append("-- Hermz & D — Movie Rankings Import")
    lines.append(f"-- Films: {len(unique_films)} | Individual: {len(individual_rankings)} | Combined: {len(combined_rankings)}")
    lines.append("-- Run AFTER movie_schema_update.sql")
    lines.append("-- ============================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # Films
    lines.append("-- ── Films ──────────────────────────────────────────────────")
    for f in unique_films:
        acc = f['acclaim_score']
        noms = f.get('oscar_nominations') or 0
        wins = f.get('oscar_wins') or 0
        lines.append(
            f"INSERT INTO public.films "
            f"(id, title, release_year, director, "
            f"actor_1, actor_2, actor_3, actor_4, "
            f"custom_genre_1, custom_genre_2, "
            f"acclaim_score, oscar_nominations, oscar_wins, "
            f"won_best_picture, won_best_director, "
            f"won_best_actor, won_best_actress, "
            f"won_best_supp_actor, won_best_supp_actress, "
            f"won_screenplay, won_cinematography, won_production_design, "
            f"afi_top100_rank, afi_comedies_rank, imdb_top250_rank, "
            f"nyt_2000s_rank, sight_sound_2022_rank, variety_comedies_rank, "
            f"national_film_registry) VALUES ("
            f"{f['id']}, {sql_str(f['title'])}, {sql_int(f['release_year'])}, "
            f"{sql_str(f['director'])}, "
            f"{sql_str(f['actor_1'])}, {sql_str(f['actor_2'])}, "
            f"{sql_str(f['actor_3'])}, {sql_str(f['actor_4'])}, "
            f"{sql_str(f.get('custom_genre_1'))}, {sql_str(f.get('custom_genre_2'))}, "
            f"{'NULL' if acc is None else acc}, {noms}, {wins}, "
            f"{sql_bool(f['won_best_picture'])}, {sql_bool(f['won_best_director'])}, "
            f"{sql_bool(f['won_best_actor'])}, {sql_bool(f['won_best_actress'])}, "
            f"{sql_bool(f['won_best_supp_actor'])}, {sql_bool(f['won_best_supp_actress'])}, "
            f"{sql_bool(f['won_screenplay'])}, {sql_bool(f['won_cinematography'])}, "
            f"{sql_bool(f['won_production_design'])}, "
            f"{sql_int(f['afi_top100_rank'])}, {sql_int(f['afi_comedies_rank'])}, "
            f"{sql_int(f['imdb_top250_rank'])}, "
            f"{sql_int(f['nyt_2000s_rank'])}, {sql_int(f['sight_sound_2022_rank'])}, "
            f"{sql_int(f['variety_comedies_rank'])}, "
            f"{sql_bool(f['national_film_registry'])});"
        )

    lines.append("")
    lines.append(f"SELECT setval('public.films_id_seq', {film_counter[0]}, false);")
    lines.append("")

    # Individual rankings
    lines.append("-- ── Individual Rankings ────────────────────────────────────")
    lines.append(
        "INSERT INTO public.individual_rankings "
        "(film_id, event_id, user_id, rank, total_score, "
        "score_lead_performance, score_supp_performance, "
        "score_plot, score_dialogue, score_screenplay, "
        "score_direction, score_cinematography, score_production_design, "
        "score_influence, score_acclaim, score_personal_impact, "
        "tb_tens, tb_nines, tb_eights) VALUES"
    )
    ir_rows = []
    for r in individual_rankings:
        ir_rows.append(
            f"  ({r['film_id']}, "
            f"(SELECT id FROM public.ranking_events WHERE year={r['event_year']}), "
            f"(SELECT id FROM public.profiles WHERE username={sql_str(r['username'])}), "
            f"{r['rank']}, {sql_int(r['total_score'])}, "
            f"{sql_score(r['score_lead_performance'])}, {sql_score(r['score_supp_performance'])}, "
            f"{sql_score(r['score_plot'])}, {sql_score(r['score_dialogue'])}, "
            f"{sql_score(r['score_screenplay'])}, "
            f"{sql_score(r['score_direction'])}, {sql_score(r['score_cinematography'])}, "
            f"{sql_score(r['score_production_design'])}, "
            f"{sql_score(r['score_influence'])}, {sql_score(r['score_acclaim'])}, "
            f"{sql_score(r['score_personal_impact'], max_=20)}, "
            f"{sql_int(r['tb_tens'])}, {sql_int(r['tb_nines'])}, {sql_int(r['tb_eights'])})"
        )
    lines.append(",\n".join(ir_rows) + ";")
    lines.append("")

    # Combined rankings
    lines.append("-- ── Combined Rankings ──────────────────────────────────────")
    lines.append(
        "INSERT INTO public.combined_rankings "
        "(film_id, event_id, combined_rank, dustin_rank, matt_rank, avg_rank, "
        "dustin_score, matt_score, total_score, "
        "dustin_impact, matt_impact, total_impact, total_tens) VALUES"
    )
    cr_rows = []
    for r in combined_rankings:
        try:
            avg = str(float(r['avg_rank'])) if r['avg_rank'] is not None else "NULL"
        except (TypeError, ValueError):
            avg = "NULL"
        cr_rows.append(
            f"  ({r['film_id']}, "
            f"(SELECT id FROM public.ranking_events WHERE year={r['event_year']}), "
            f"{r['combined_rank']}, "
            f"{sql_int(r['dustin_rank'])}, {sql_int(r['matt_rank'])}, {avg}, "
            f"{sql_int(r['dustin_score'])}, {sql_int(r['matt_score'])}, "
            f"{sql_int(r['total_score'])}, "
            f"{sql_int(r['dustin_impact'])}, {sql_int(r['matt_impact'])}, "
            f"{sql_int(r['total_impact'])}, {sql_int(r['total_tens'])})"
        )
    lines.append(",\n".join(cr_rows) + ";")
    lines.append("")
    lines.append("COMMIT;")
    lines.append("")
    lines.append("-- Validation:")
    lines.append("-- SELECT COUNT(*) FROM public.films;")
    lines.append("-- SELECT COUNT(*) FROM public.individual_rankings;")
    lines.append("-- SELECT COUNT(*) FROM public.combined_rankings;")

    OUTPUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n✓  Wrote {OUTPUT}  ({OUTPUT.stat().st_size:,} bytes)")

    # Breakdown
    from collections import Counter
    print("\nIndividual rankings by event + person:")
    for k, v in sorted(Counter((r['event_year'], r['username']) for r in individual_rankings).items()):
        print(f"  {k[0]} {k[1]:6}: {v}")
    print("\nCombined rankings by event:")
    for k, v in sorted(Counter(r['event_year'] for r in combined_rankings).items()):
        print(f"  {k}: {v}")
    print(f"\nFilms with director:       {sum(1 for f in unique_films if f['director'])}/{len(unique_films)}")
    print(f"Films with acclaim score:  {sum(1 for f in unique_films if f['acclaim_score'])}/{len(unique_films)}")
    print(f"Films with NULL year:      {len(null_year)}")


if __name__ == "__main__":
    main()
